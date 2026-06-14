"""
Data Export Routes
CSV, Excel, and JSON export for third-party integration.
"""

import csv
import io
from datetime import date, datetime, timedelta, timezone

import psycopg2
import psycopg2.extras
from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse

router = APIRouter()

DB_DSN = "postgresql://agritech:agritech@localhost:5432/agritech"


def get_db():
    return psycopg2.connect(DB_DSN, cursor_factory=psycopg2.extras.RealDictCursor)


# ============================================================
# CSV Export
# ============================================================

@router.get("/csv")
async def export_csv(
    farm_ids: str = Query(default=None, description="Comma-separated farm UUIDs"),
    region_id: int = Query(default=None),
    sensor_types: str = Query(default=None, description="Comma-separated sensor types"),
    start_date: date = Query(default=None),
    end_date: date = Query(default=None),
    granularity: str = Query("daily", regex="^(raw|hourly|daily)$"),
):
    """
    Export sensor data as CSV.
    Supports filtering by farm, region, sensor type, and date range.
    """
    if not start_date:
        start_date = date.today() - timedelta(days=30)
    if not end_date:
        end_date = date.today()

    # Parse filters
    farm_id_list = farm_ids.split(",") if farm_ids else None
    sensor_type_list = sensor_types.split(",") if sensor_types else None

    # Choose source table
    if granularity == "raw":
        table = "sensor_readings"
        time_col = "sr.time"
        value_cols = "sr.reading_value AS value"
    elif granularity == "hourly":
        table = "readings_hourly"
        time_col = "bucket AS time"
        value_cols = "avg_value, min_value, max_value, sample_count"
    else:  # daily
        table = "readings_daily"
        time_col = "bucket AS time"
        value_cols = "avg_value, min_value, max_value, sample_count"

    conn = get_db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            # Build query
            if granularity == "raw":
                query = f"""
                    SELECT {time_col}, sr.farm_id, sr.zone_id, s.sensor_type,
                           {value_cols}, sr.quality_flag, sr.battery_level
                    FROM sensor_readings sr
                    JOIN sensors s ON s.sensor_id = sr.sensor_id
                    JOIN zones z ON z.zone_id = sr.zone_id
                    WHERE sr.time >= %s AND sr.time < %s
                """
            else:
                query = f"""
                    SELECT {time_col}, farm_id, zone_id, sensor_type,
                           {value_cols}
                    FROM {table}
                    WHERE bucket >= %s AND bucket < %s
                """

            params = [start_date, end_date + timedelta(days=1)]

            if farm_id_list:
                id_col = "sr.farm_id" if granularity == "raw" else "farm_id"
                placeholders = ",".join(["%s"] * len(farm_id_list))
                query += f" AND {id_col} IN ({placeholders})"
                params.extend(farm_id_list)

            if region_id:
                query += f" AND {'f.region_id' if granularity == 'raw' else 'farm_id IN (SELECT farm_id FROM farms WHERE region_id = %s)'}"
                if granularity != "raw":
                    params.append(region_id)

            if sensor_type_list:
                type_col = "s.sensor_type" if granularity == "raw" else "sensor_type"
                placeholders = ",".join(["%s"] * len(sensor_type_list))
                query += f" AND {type_col} IN ({placeholders})"
                params.extend(sensor_type_list)

            time_order = "sr.time" if granularity == "raw" else "bucket"
            query += f" ORDER BY {time_order} LIMIT 1000000"

            cur.execute(query, params)

            # Stream CSV
            def generate_csv():
                output = io.StringIO()
                writer = csv.writer(output)

                # Header
                if cur.description:
                    writer.writerow([desc[0] for desc in cur.description])
                    yield output.getvalue()
                    output.seek(0)
                    output.truncate(0)

                # Rows
                for row in cur:
                    writer.writerow([str(v) if v is not None else "" for v in row.values()])
                    yield output.getvalue()
                    output.seek(0)
                    output.truncate(0)

            filename = f"agritech_export_{start_date}_{end_date}_{granularity}.csv"
            return StreamingResponse(
                generate_csv(),
                media_type="text/csv",
                headers={"Content-Disposition": f"attachment; filename={filename}"},
            )
    finally:
        conn.close()


# ============================================================
# Excel Export
# ============================================================

@router.get("/excel")
async def export_excel(
    farm_ids: str = Query(default=None),
    start_date: date = Query(default=None),
    end_date: Date = Query(default=None),
    include_summary: bool = Query(True),
):
    """
    Export data as Excel workbook with multiple sheets.
    """
    if not start_date:
        start_date = date.today() - timedelta(days=30)
    if not end_date:
        end_date = date.today()

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "openpyxl not installed. Install with: pip install openpyxl")

    farm_id_list = farm_ids.split(",") if farm_ids else None

    conn = get_db()
    try:
        wb = openpyxl.Workbook()

        # ---- Sheet 1: Daily Averages ----
        ws1 = wb.active
        ws1.title = "Daily Averages"

        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            query = """
                SELECT rd.bucket AS date, f.name AS farm_name, z.name AS zone_name,
                       rd.sensor_type, rd.avg_value, rd.min_value, rd.max_value, rd.sample_count
                FROM readings_daily rd
                JOIN zones z ON z.zone_id = rd.zone_id
                JOIN farms f ON f.farm_id = rd.farm_id
                WHERE rd.bucket >= %s AND rd.bucket < %s
            """
            params = [start_date, end_date + timedelta(days=1)]

            if farm_id_list:
                placeholders = ",".join(["%s"] * len(farm_id_list))
                query += f" AND rd.farm_id IN ({placeholders})"
                params.extend(farm_id_list)

            query += " ORDER BY rd.bucket, f.name, z.name, rd.sensor_type"
            cur.execute(query, params)
            rows = cur.fetchall()

            # Header
            headers = ["Date", "Farm", "Zone", "Sensor Type", "Average", "Min", "Max", "Samples"]
            header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)

            for col, header in enumerate(headers, 1):
                cell = ws1.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font

            # Data
            for row_idx, row in enumerate(rows, 2):
                for col_idx, key in enumerate(row.keys(), 1):
                    value = row[key]
                    if hasattr(value, "isoformat"):
                        value = value.isoformat()
                    elif isinstance(value, float):
                        value = round(value, 2)
                    ws1.cell(row=row_idx, column=col_idx, value=value)

            # Auto-fit columns
            for col in range(1, len(headers) + 1):
                ws1.column_dimensions[get_column_letter(col)].width = 15

        # ---- Sheet 2: Farm Summary ----
        if include_summary:
            ws2 = wb.create_sheet("Farm Summary")
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("""
                    SELECT f.name AS farm_name, r.name AS region_name,
                           f.area_hectares, f.coffee_variety,
                           COUNT(DISTINCT z.zone_id) AS zones,
                           COUNT(DISTINCT s.sensor_id) AS sensors,
                           MAX(l.last_reading_at) AS last_data
                    FROM farms f
                    JOIN regions r ON r.region_id = f.region_id
                    JOIN zones z ON z.farm_id = f.farm_id
                    JOIN sensors s ON s.zone_id = z.zone_id
                    LEFT JOIN latest_readings l ON l.sensor_id = s.sensor_id
                    WHERE f.is_active = TRUE
                    GROUP BY f.name, r.name, f.area_hectares, f.coffee_variety
                    ORDER BY f.name
                """)
                summary_rows = cur.fetchall()

                headers2 = ["Farm", "Region", "Area (ha)", "Variety", "Zones", "Sensors", "Last Data"]
                for col, header in enumerate(headers2, 1):
                    cell = ws2.cell(row=1, column=col, value=header)
                    cell.fill = header_fill
                    cell.font = header_font

                for row_idx, row in enumerate(summary_rows, 2):
                    for col_idx, value in enumerate(row.values(), 1):
                        if hasattr(value, "isoformat"):
                            value = value.isoformat()
                        ws2.cell(row=row_idx, column=col_idx, value=value)

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"agritech_report_{start_date}_{end_date}.xlsx"
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    finally:
        conn.close()


# ============================================================
# JSON API (for third-party integration)
# ============================================================

@router.get("/json")
async def export_json(
    farm_id: str = Query(...),
    sensor_type: SensorType = Query(...),
    start_date: date = Query(default=None),
    end_date: date = Query(default=None),
):
    """
    JSON export for third-party integration.
    Returns data in a standardized format compatible with common AgriTech APIs.
    """
    if not start_date:
        start_date = date.today() - timedelta(days=30)
    if not end_date:
        end_date = date.today()

    conn = get_db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("""
                SELECT bucket AS timestamp, farm_id, zone_id, sensor_type,
                       avg_value AS value, min_value, max_value, sample_count
                FROM readings_daily
                WHERE farm_id = %s AND sensor_type = %s
                    AND bucket >= %s AND bucket < %s
                ORDER BY bucket
            """, (farm_id, sensor_type.value, start_date, end_date + timedelta(days=1)))

            rows = cur.fetchall()

            # Convert types
            for row in rows:
                for k, v in row.items():
                    if hasattr(v, "isoformat"):
                        row[k] = v.isoformat()
                    elif isinstance(v, float):
                        row[k] = round(v, 2)

            return {
                "metadata": {
                    "farm_id": farm_id,
                    "sensor_type": sensor_type.value,
                    "period": {"start": str(start_date), "end": str(end_date)},
                    "granularity": "daily",
                    "count": len(rows),
                },
                "data": rows,
            }
    finally:
        conn.close()
